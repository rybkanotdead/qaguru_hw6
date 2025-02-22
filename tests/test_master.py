import csv
from pypdf import PdfReader
from zipfile import ZipFile
from tests.conftest import arhive_path
from openpyxl import load_workbook

def test_xlsx():
    with ZipFile(arhive_path) as zip_file:
        with zip_file.open("file_example_XLSX_50.xlsx") as xlsx_file:
            excel_file = load_workbook(xlsx_file)
            excel_sheet = excel_file.active

            assert  excel_sheet.max_row == 51
            assert excel_sheet.max_column == 8

def tests_csv():
    with ZipFile(arhive_path) as zip_file:
        with zip_file.open("file_example_csv_10.csv") as csv_file:
            csv_file = csv.reader(csv_file.read().decode("utf-8").splitlines(), delimiter=";")
            csv_rows = list(csv_file)

        assert csv_rows[0] == ["0", "First Name", "Last Name", "Gender", "Country", "Age", "Date", "Id"]

def tests_pdf():
    with ZipFile(arhive_path) as zip_file:
        with zip_file.open("file_example_pdf_10.pdf") as pdf_file:
            file_pdf = PdfReader(pdf_file)
            text = file_pdf.pages[0].extract_text()

            assert "Sheet1" in text